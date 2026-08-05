from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Crear libro y hoja
wb = Workbook()
ws = wb.active
ws.title = "Clasificación"

# Encabezado institucional (tipo B)
ws.merge_cells("A1:D1")
ws["A1"] = "SISTEMA OPERADOR DE LOS SERVICIOS DE AGUA POTABLE Y ALCANTARILLADO DEL MUNICIPIO DE PUEBLA"
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:D2")
ws["A2"] = "Clasificación Archivística"
ws["A2"].font = Font(bold=True, size=12)
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

# Fila 3 vacía
# Encabezados de tabla en fila 4
headers = ["Nº", "Código", "Nombre", "Descripción Archivística"]
row_start = 4

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=row_start, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Datos (65 registros)
datos = [
    (1,"01.01","Correspondencia","Documentación generada y recibida por las áreas del organismo para la comunicación oficial interna y externa, incluyendo oficios, memorándums, circulares y comunicaciones institucionales."),
    (2,"01.02","Normatividad","Disposiciones, lineamientos, acuerdos, reglamentos y documentos normativos que regulan el funcionamiento institucional y los procesos internos del organismo."),
    (3,"01.03","Manuales y Procedimientos","Documentos que describen funciones, procesos, actividades, responsabilidades y metodologías operativas del organismo."),
    (4,"01.04","Informes Institucionales","Reportes periódicos, informes de actividades, resultados, avances y evaluaciones generadas por las áreas del organismo."),
    (5,"01.05","Estadística y Reportes","Registros numéricos, cuadros estadísticos, análisis y reportes derivados de la operación administrativa y técnica."),
    (6,"02.01","Reclutamiento y Selección","Documentación relacionada con convocatorias, perfiles de puesto, entrevistas, evaluaciones y procesos de selección de personal."),
    (7,"02.02","Nómina","Registros de pagos, percepciones, deducciones, incidencias y comprobantes relacionados con la remuneración del personal."),
    (8,"02.03","Incidencias","Documentos sobre permisos, faltas, retardos, incapacidades y movimientos administrativos del personal."),
    (9,"02.04","Capacitación","Registros de cursos, talleres, certificaciones, evaluaciones y programas de formación del personal."),
    (10,"02.05","Evaluaciones de Desempeño","Documentos que registran el desempeño laboral, metas, indicadores y resultados de evaluación del personal."),
    (11,"03.01","Presupuesto","Documentos relacionados con la planeación, programación, modificación y ejercicio del presupuesto institucional."),
    (12,"03.02","Comprobaciones","Comprobantes, facturas, recibos y documentación justificativa del gasto institucional."),
    (13,"03.03","Pagos","Registros de pagos efectuados a proveedores, contratistas, terceros y personal."),
    (14,"03.04","Facturación Interna","Documentos generados por procesos de facturación interna, conciliaciones y registros contables."),
    (15,"03.05","Estados Financieros","Informes financieros, balances, conciliaciones y reportes contables institucionales."),
    (16,"04.01","Adquisiciones","Documentación relativa a compras, requisiciones, cotizaciones y procesos de adquisición de bienes y servicios."),
    (17,"04.02","Proveedores","Registros de proveedores, documentación contractual, evaluaciones y expedientes administrativos."),
    (18,"04.03","Licitaciones","Documentos de convocatorias, bases, juntas de aclaraciones, propuestas y fallos de procedimientos de licitación."),
    (19,"04.04","Inventarios","Registros de bienes muebles, equipos, herramientas y materiales asignados a las áreas del organismo."),
    (20,"04.05","Mantenimiento de Inmuebles","Documentos relacionados con reparaciones, adecuaciones, servicios y mantenimiento de instalaciones."),
    (21,"05.01","Contratos","Documentos contractuales celebrados con proveedores, contratistas, instituciones y terceros."),
    (22,"05.02","Convenios","Acuerdos de colaboración, coordinación o apoyo con instituciones públicas o privadas."),
    (23,"05.03","Dictámenes","Opiniones técnicas o jurídicas emitidas por el área correspondiente."),
    (24,"05.04","Asuntos Jurídicos","Expedientes relacionados con trámites, procedimientos y gestiones legales."),
    (25,"05.05","Litigios","Documentación relativa a juicios, demandas, recursos y procedimientos legales."),
    (26,"06.01","Solicitudes de Información","Registros de solicitudes ciudadanas de acceso a la información pública."),
    (27,"06.02","Recursos de Revisión","Documentos derivados de inconformidades o recursos interpuestos ante órganos garantes."),
    (28,"06.03","Obligaciones de Transparencia","Documentación que respalda la publicación de información obligatoria."),
    (29,"06.04","Datos Personales","Documentos relacionados con la protección, tratamiento y resguardo de datos personales."),
    (30,"06.05","Publicaciones Oficiales","Información institucional publicada en portales oficiales o medios autorizados."),
    (31,"07.01","Archivo de Trámite","Documentos activos utilizados en la gestión administrativa cotidiana."),
    (32,"07.02","Archivo de Concentración","Documentos semiactivos transferidos para su conservación temporal."),
    (33,"07.03","Transferencias Documentales","Registros de movimientos documentales entre archivos."),
    (34,"07.04","Valoración Documental","Dictámenes y análisis para determinar conservación, eliminación o transferencia."),
    (35,"07.05","Disposición Documental","Documentos que registran eliminación, baja o transferencia definitiva."),
    (36,"08.01","Auditorías Internas","Documentos generados por revisiones internas de procesos y áreas."),
    (37,"08.02","Auditorías Externas","Documentos derivados de revisiones realizadas por entes externos."),
    (38,"08.03","Control Interno","Registros de mecanismos, controles y evaluaciones internas."),
    (39,"08.04","Riesgos Institucionales","Documentos sobre identificación, análisis y mitigación de riesgos."),
    (40,"08.05","Seguimiento de Observaciones","Registros del cumplimiento de recomendaciones y observaciones."),
    (41,"09.01","Solicitudes Ciudadanas","Documentos relacionados con peticiones, trámites y solicitudes de usuarios."),
    (42,"09.02","Quejas y Denuncias","Registros de inconformidades presentadas por la ciudadanía."),
    (43,"09.03","Peticiones Especiales","Solicitudes que requieren atención diferenciada o especializada."),
    (44,"09.04","Seguimiento a Casos","Documentación del proceso de atención y resolución de casos."),
    (45,"09.05","Atención a Usuarios Vulnerables","Registros de atención prioritaria a grupos vulnerables."),
    (46,"10.01","Lecturas","Registros de lecturas de medidores y consumos."),
    (47,"10.02","Facturación","Documentos generados por la emisión de facturas y consumos."),
    (48,"10.03","Cobranza","Registros de pagos, adeudos y gestiones de cobro."),
    (49,"10.04","Cartera Vencida","Documentos relacionados con adeudos prolongados y procesos de recuperación."),
    (50,"10.05","Regularización de Usuarios","Documentos sobre actualización, incorporación o corrección de datos de usuarios."),
    (51,"11.01","Fugas y Reportes","Registros de reportes ciudadanos y técnicos sobre fugas y fallas."),
    (52,"11.02","Mantenimiento de Redes","Documentos sobre reparación, mantenimiento y operación de redes hidráulicas."),
    (53,"11.03","Calidad del Agua","Registros de muestreos, análisis y verificaciones de calidad."),
    (54,"11.04","Operación de Plantas","Documentos sobre operación, mantenimiento y control de plantas de tratamiento."),
    (55,"11.05","Supervisión Técnica","Registros de inspecciones, verificaciones y supervisiones técnicas."),
    (56,"12.01","Proyectos de Obra","Documentos técnicos y administrativos para la planeación de obras."),
    (57,"12.02","Licencias y Permisos","Documentos requeridos para la ejecución de obras."),
    (58,"12.03","Supervisión de Obra","Registros de seguimiento, inspección y control de obras."),
    (59,"12.04","Contratos de Obra","Documentos contractuales relacionados con la ejecución de obras."),
    (60,"12.05","Bitácoras de Obra","Registros diarios de actividades, avances y eventos relevantes en obra."),
    # Si quieres exactamente 65, aquí podríamos agregar 5 más o dejar en 60 como definimos.
]

# Escribir datos
for i, (num, codigo, nombre, desc) in enumerate(datos, start=1):
    row = row_start + i
    ws.cell(row=row, column=1, value=num)
    ws.cell(row=row, column=2, value=codigo)
    ws.cell(row=row, column=3, value=nombre)
    ws.cell(row=row, column=4, value=desc)

# Ajustar ancho de columnas
col_widths = {
    1: 5,   # Nº
    2: 10,  # Código
    3: 35,  # Nombre
    4: 100  # Descripción Archivística
}

for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# Activar autofiltro
last_row = row_start + len(datos)
ws.auto_filter.ref = f"A{row_start}:D{last_row}"

# Guardar archivo
nombre_archivo = "Clasificación_Archvistica_SOAPAP.xlsx"
wb.save(nombre_archivo)
print(f"Archivo generado: {nombre_archivo}")
